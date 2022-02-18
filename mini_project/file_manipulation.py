import os
import log
import openpyxl
import datetime
import shutil

def find_file():
    try:
        path_array = []
        for file in os.listdir('./mini_project/export_files'):
            if file.startswith("expedia_report_monthly_"):
                path = './mini_project/export_files/' + file
                log.log("File loaded successfully")
                path_array.append(path)
                src_folder = "/mini_project/export_files//"
                if os.path.exists(src_folder):
                    shutil.move(path, src_folder)
            else:
                file_path = './mini_project/export_files/' + file
                error_file(file_path)

        if len(path_array) > 0:
            return path_array
    except:
        return log.log("Error while loading file")

def read_file(file):
    try:
        if not file:
            log.log("No file found, using the naming convention expedia_report_monthly_MONTH_YEAR.xlsx")
        else:
            try:
                wb = openpyxl.load_workbook(file)
                return wb
            except:
                log.log("Data was not able to load")
                error_file(file)
    except:
        log.log("Error while reading File")
        error_file(file)


def find_row(ws, search):
    for x in range(1, ws.max_column + 1):
        for y in range(1, ws.max_row + 1):
            current_cell = ws.cell(row = y, column = x).value
            if isinstance(current_cell, str) and isinstance(search, str):
                if search in current_cell:
                    return y
            elif isinstance(current_cell, datetime.date) and isinstance(search, datetime.datetime):
                if current_cell.year == search.year and current_cell.month == search.month:
                    return y
            else:
                continue
    log.log("No row found matching the date/field information")

def find_column(ws, name):
    for x in range(1, ws.max_column + 1):
        for y in range(1, ws.max_row + 1):
            if ws.cell(row = y, column = x).value == name:
                return x
    log.log("No Column found matching the date/field information")
                
def capture_data(ws, my_row):
    calls = ws.cell(row = my_row, column = find_column(ws, 'Calls Offered')).value
    abandon = ws.cell(row = my_row, column = find_column(ws, ' Abandon after 30s')).value
    fcr = ws.cell(row = my_row, column = find_column(ws, 'FCR')).value
    dsat = ws.cell(row = my_row, column = find_column(ws, 'DSAT ')).value
    csat = ws.cell(row = my_row, column = find_column(ws, 'CSAT ')).value

    if calls and abandon and fcr and dsat and csat:
        data = {
        'Calls Offered': calls,
        'Abandon after 30s': "{:.2%}".format(abandon),
        'FCR': "{:.2%}".format(fcr),
        'DSAT': "{:.2%}".format(dsat),
        'CSAT': "{:.2%}".format(csat)
        }
        return data
    else:
        return -1

def get_data(file, date):
    if(date == -1):
        log.log("File naming incorrect")
        error_file(file)
    else:
        log.log("Date of Data extracted correctly")
        wb = read_file(file)
        data_date = datetime.datetime(int(date[0]), date[1], 1)
        ws = wb.active
        my_row = find_row(ws, data_date)

        if my_row:
            data = capture_data(ws, my_row)

        if not my_row or data == -1:
            log.log("Fields not found, file moved to error_files folder")
            error_file(file)
            return -1
        
        else:
            data_month = '' + date[2]
            log.log("Data for {}: \n {}".format(data_month, data))
            return data

def get_data_from_sheet(file, sheet, date):
    if(date == -1):
        log.log("File naming incorrect")
        error_file(file)
    else:
        log.log("Date of Data extracted correctly")
        data_date = datetime.datetime(int(date[0]), date[1], 1)
        wb = read_file(file)
        ws2 = wb[sheet]
        my_col = find_column(ws2, data_date)

        if not my_col:
            my_col = find_column(ws2, date[2])
    
        if not my_col:
            error_file(file)
            log.log("Column not found, file moved to Error folder")
        # Check Promoters
        my_row = find_row(ws2, 'Promoters')
        promoters = check_score('Promoters', ws2.cell(row = my_row, column = my_col).value)

        # Check Passives
        my_row = find_row(ws2, 'Passives')
        passives = check_score('Passives', ws2.cell(row = my_row, column = my_col).value)
        
        # Check Decractors
        my_row = find_row(ws2, 'Dectractors')
        dectractors = check_score('Dectractors', ws2.cell(row = my_row, column = my_col).value)

        if promoters and passives and dectractors:
            log.log("In Net Promoters Score: \n\tPromoters: {} \n\tPassives: {} \n\tDectractors: {}".format(promoters, passives, dectractors))
            arquive_file(file)
        
        else:
            log.log("Fields not found, file moved to error_files folder")
            error_file(file)

def check_score(category, value):
    if category == 'Promoters':
        if value > 200:
            return 'good'
        else:
            return 'bad'
    elif category == 'Passives':
        if value > 100:
            return 'good'
        else:
            return 'bad'
    elif category == 'Dectractors':
        if value > 100:
            return 'good'
        else:
            return 'bad'
    else:
        log.log("Could not find the field {}".format(category))

def arquive_file(file):
    src_folder = "./mini_project/arquived_files//"
    if os.path.exists(src_folder):
        shutil.move(file, src_folder)

def error_file(error_file):
    error_f = error_file.split('./mini_project/export_files/')[1]
    src_folder = "./mini_project/error_files//"
    log.log("File being moved to Error folder {}".format(error_f))

    if len(os.listdir('./mini_project/error_files')) > 0:
        for file in os.listdir('./mini_project/error_files'):
            if file == error_file.split('./mini_project/export_files/')[1]:
                break
            else:
                if os.path.exists(src_folder):
                    shutil.move(error_file, src_folder)
                    log.log("File moved successfully into Error_Files folder")
    else:    
        if os.path.exists(src_folder):
            shutil.move(error_file, src_folder)
            log.log("File moved successfully into Error_Files folder")

# get_data_from_sheet('./mini_project/export_files/expedia_report_monthly_march_2018.xlsx', 'VOC Rolling MoM', ['2018', 1, 'March'])
# wb = read_file('./mini_project/export_files/expedia_report_monthly_january_2018.xlsx')
# ws = wb.active
# date = ['2018', 1, 'January']
# data_date = datetime.datetime(int(date[0]), date[1], 1)
# find_row(ws, data_date)