import openpyxl
import os
import datetime
import logging

logging.basicConfig(filename='test.log', level = logging.DEBUG)

def find_file():
    path = -1     
    for file in os.listdir('./Day 5'):
        if file.startswith("expedia_report_monthly_"):
            path = './Day 5/' + file
            logging.info("File loaded successfully")
    return path

def month_year(file):
    date = file.split("expedia_report_monthly_",1)[1]
    year = int(date[-9:-5])
    month = date[0:3]
    month_name = date[0:-10]

    if(month == 'jan'):
        return [year, 1, month_name]
    elif(month == 'feb'):
        return [year, 2, month_name]
    elif(month == 'mar'):
        return [year, 3, month_name]
    elif(month == 'apr'):
        return [year, 4, month_name]
    elif(month == 'may'):
        return [year, 5, month_name]
    elif(month == 'jun'):
        return [year, 6, month_name]
    elif(month == 'jul'):
        return [year, 7, month_name]
    elif(month == 'aug'):
        return [year, 8, month_name]
    elif(month == 'sep'):
        return [year, 9, month_name]
    elif(month == 'oct'):
        return [year, 10, month_name]
    elif(month == 'nov'):
        return [year, 11, month_name]
    elif(month == 'dec'):
        return [year, 12, month_name]
    else:
        return -1


def read_file():
    path = find_file()
    if path == -1:
        logging.debug("No file found, using the naming convention expedia_report_monthly_MONTH_YEAR.xlsx")
    
    else:
        date = month_year(path)

        if(date == -1):
            logging.debug("File naming incorrect")
        else:
            logging.info("Date of Data extracted correctly")
            data_date = datetime.datetime(date[0], date[1], 1)
            
            wb = openpyxl.load_workbook(path)
            ws = wb.active

            data=[]
            my_col = None
            my_row = None
        
            for col in ws.iter_cols(min_row=0, max_row=ws.max_row + 1):
                for cell in col:
                    if isinstance(cell.value, datetime.date) and cell.value.year == data_date.year and cell.value.month == data_date.month:
                        my_col, my_row = cell.column, cell.row
                        logging.info("Column and Row of data found")
        
            for x in range(2, my_col + 6):
                data.append(ws.cell(row = my_row, column = x).value)

            data_dict = {
                'Calls Offered: ': data[0],
                'Abandon after 30s : ': data[1],
                'FCR : ': data[2],
                'DSAT : ': data[3],
                'CSAT : ': data[4]
            }

            data_month = '' + date[2]
            logging.info("Data for {}: \n {}".format(data_month.capitalize(), data_dict))
            return data_dict

read_file()
